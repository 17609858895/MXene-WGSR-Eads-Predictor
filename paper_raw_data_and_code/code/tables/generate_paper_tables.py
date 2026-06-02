from __future__ import annotations

import importlib.util
import json
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from scipy.stats import spearmanr, wilcoxon


ROOT = Path(__file__).resolve().parents[2]
FIG_SCRIPT = ROOT / "code" / "figures" / "generate_all_figures.py"
OUT = ROOT / "outputs" / "tables"
OUT.mkdir(parents=True, exist_ok=True)


def load_figure_module():
    spec = importlib.util.spec_from_file_location("figures", FIG_SCRIPT)
    module = importlib.util.module_from_spec(spec)
    assert spec.loader is not None
    spec.loader.exec_module(module)
    return module


def missing_rate(series: pd.Series) -> float:
    return float(series.replace("-", np.nan).isna().mean() * 100)


def fmt_float(x, digits=3):
    if pd.isna(x):
        return ""
    return round(float(x), digits)


def preprocessing_note(feature: str, category: str) -> str:
    if category in {"Target", "Secondary target"}:
        return "Used as label; rows with missing labels are excluded from the corresponding task."
    if feature in {"Stacking", "M1", "M2", "X", "T1_type", "Mol", "Dopant"}:
        return "Missing markers converted to NaN; categorical variables are imputed and one-hot encoded when used for modelling."
    return "Missing markers converted to NaN; numeric variables are median-imputed in the modelling pipeline."


FEATURE_META = {
    "a_ang": ("Structural", "Lattice constant", "Angstrom", "In-plane lattice constant of MXene."),
    "Stacking": ("Structural", "Stacking sequence", "-", "Layer stacking mode such as ABA or ABC."),
    "N_Layers": ("Structural", "Number of layers", "-", "Number of atomic layers in the MXene slab."),
    "Layer_dist_MX": ("Structural", "M-X layer distance", "Angstrom", "Distance between M and X layers."),
    "Layer_dist_MM": ("Structural", "M-M layer distance", "Angstrom", "Distance between adjacent metal layers."),
    "Length_MX": ("Structural", "M-X bond length", "Angstrom", "Bond length between transition metal and C/N."),
    "Length_MT1": ("Structural", "M-T bond length", "Angstrom", "Bond length between metal and first termination."),
    "E_Form": ("Electronic/thermodynamic", "Formation energy", "eV", "DFT formation energy of the MXene structure."),
    "Bader_M1": ("Electronic", "Bader charge of M1", "e", "Bader charge on primary metal site."),
    "Bader_M2": ("Electronic", "Bader charge of M2", "e", "Bader charge on secondary metal site, if present."),
    "Bader_X": ("Electronic", "Bader charge of X", "e", "Bader charge on carbon or nitrogen site."),
    "Bader_T1": ("Electronic", "Bader charge of T1", "e", "Bader charge on surface termination."),
    "Band_gap_PBE_ev": ("Electronic", "PBE band gap", "eV", "Band gap calculated using PBE functional."),
    "M1": ("Composition", "Primary metal", "-", "Primary transition metal element."),
    "M2": ("Composition", "Secondary metal", "-", "Secondary metal element, if present."),
    "X": ("Composition", "X element", "-", "C or N in MXene."),
    "T1_type": ("Surface chemistry", "Termination type", "-", "Primary surface termination group."),
    "Mol": ("Adsorbate", "Adsorbate species", "-", "WGSR-relevant adsorbate species."),
    "Dopant": ("Dopant domain", "Dopant element", "-", "Dopant element in external dopants sheet."),
    "E_ads_ev": ("Target", "Adsorption energy", "eV", "Main prediction target."),
    "E_activation_ev": ("Secondary target", "Activation energy", "eV", "Small-sample mechanistic target."),
    "E_reaction_ev": ("Secondary target", "Reaction energy", "eV", "Small-sample reaction-energy target."),
}


def make_table1(wgs: pd.DataFrame, dop: pd.DataFrame) -> pd.DataFrame:
    rows = []
    for sheet, df in [("WGS", wgs), ("dopants", dop)]:
        total = len(df)
        for target in ["E_ads_ev", "E_activation_ev", "E_reaction_ev", "Reaction"]:
            valid = int(df[target].replace("-", np.nan).notna().sum()) if target in df.columns else 0
            rows.append(
                {
                    "Dataset": sheet,
                    "Rows": total,
                    "Variable": target,
                    "Role in paper": {
                        "E_ads_ev": "Main prediction target",
                        "E_activation_ev": "Small-sample mechanistic analysis",
                        "E_reaction_ev": "Small-sample thermodynamic analysis",
                        "Reaction": "Reaction-path grouping",
                    }[target],
                    "Valid labels": valid,
                    "Missing labels": total - valid,
                    "Missing rate (%)": round((total - valid) / total * 100, 2),
                }
            )
    return pd.DataFrame(rows)


def make_table2(wgs: pd.DataFrame, dop: pd.DataFrame) -> pd.DataFrame:
    combined = pd.concat([wgs.assign(Dataset="WGS"), dop.assign(Dataset="dopants")], ignore_index=True)
    rows = []
    for feature, (category, display, unit, definition) in FEATURE_META.items():
        if feature not in combined.columns:
            continue
        s = combined[feature].replace("-", np.nan)
        numeric = pd.to_numeric(s, errors="coerce")
        is_num = numeric.notna().sum() >= max(10, int(s.notna().sum() * 0.7))
        row = {
            "Category": category,
            "Feature": feature,
            "Display name": display,
            "Unit": unit,
            "Definition": definition,
            "Preprocessing": preprocessing_note(feature, category),
            "Valid count": int(s.notna().sum()),
            "Missing rate (%)": round(s.isna().mean() * 100, 2),
            "Unique values": int(s.nunique(dropna=True)),
        }
        if is_num:
            row.update(
                {
                    "Mean": fmt_float(numeric.mean()),
                    "Std": fmt_float(numeric.std()),
                    "Min": fmt_float(numeric.min()),
                    "Max": fmt_float(numeric.max()),
                }
            )
        else:
            row.update({"Mean": "", "Std": "", "Min": "", "Max": ""})
        rows.append(row)
    return pd.DataFrame(rows)


def make_table3(context: dict) -> pd.DataFrame:
    df = context["metrics"].copy()
    df.index.name = "Model"
    df = df.reset_index()
    order = ["Model", "R2", "RMSE", "MAE", "MAPE", "Train_R2", "Train_RMSE", "Train_MAE", "Train_MAPE"]
    df = df[order]
    for col in df.columns:
        if col != "Model":
            df[col] = df[col].astype(float).round(4)
    df["Notes"] = np.where(df["Model"].eq("TOPSIS_ensemble"), "Entropy-TOPSIS weighted ensemble", "Single model")
    return df


def evaluate_model_suite(fig, x_train: pd.DataFrame, x_test: pd.DataFrame, y_train: pd.Series, y_test: pd.Series) -> pd.DataFrame:
    preprocessor = fig.build_preprocessor(pd.concat([x_train, x_test], axis=0))
    models = fig.make_models(preprocessor)
    rows = []
    preds_test = {}
    preds_train = {}
    for name, pipe in models.items():
        pipe.fit(x_train, y_train)
        pred_test = pipe.predict(x_test)
        pred_train = pipe.predict(x_train)
        preds_test[name] = pred_test
        preds_train[name] = pred_train
        rows.append(
            {
                "Model": name,
                **fig.metrics(y_test, pred_test),
                **{f"Train_{k}": v for k, v in fig.metrics(y_train, pred_train).items()},
            }
        )
    metric_df = pd.DataFrame(rows).set_index("Model")
    topsis = fig.entropy_topsis(metric_df[["R2", "RMSE", "MAE", "MAPE"]])
    weights = topsis["ensemble_weight"].reindex(metric_df.index).to_numpy()
    ens_test = np.column_stack([preds_test[m] for m in metric_df.index]) @ weights
    ens_train = np.column_stack([preds_train[m] for m in metric_df.index]) @ weights
    metric_df.loc["TOPSIS_ensemble", ["R2", "RMSE", "MAE", "MAPE"]] = list(fig.metrics(y_test, ens_test).values())
    metric_df.loc["TOPSIS_ensemble", ["Train_R2", "Train_RMSE", "Train_MAE", "Train_MAPE"]] = list(fig.metrics(y_train, ens_train).values())
    return metric_df


def make_table_s1b_missingness_sensitivity(context: dict, fig, wgs_raw: pd.DataFrame, threshold: float = 0.50) -> pd.DataFrame:
    descriptor_sources = [c for c in fig.NUMERIC_BASE + fig.CATEGORICAL_BASE + ["Dopant"] if c in wgs_raw.columns]
    high_missing_sources = [
        c for c in descriptor_sources if wgs_raw[c].replace("-", np.nan).isna().mean() >= threshold
    ]
    removed_cols = [
        c
        for c in context["x_wgs"].columns
        if any(c == source or c.startswith(f"{source}_") for source in high_missing_sources)
    ]
    retained_cols = [c for c in context["x_wgs"].columns if c not in removed_cols]

    baseline = context["metrics"].copy()
    filtered = evaluate_model_suite(
        fig,
        context["x_train"][retained_cols],
        context["x_test"][retained_cols],
        context["y_train"],
        context["y_test"],
    )
    rows = []
    removed_source_text = ", ".join(high_missing_sources)
    for feature_set, df in [
        ("All descriptors", baseline),
        (f"Descriptors with WGS missingness >= {int(threshold * 100)}% removed", filtered),
    ]:
        for model, row in df.iterrows():
            base = baseline.loc[model]
            delta_r2 = float(row["R2"] - base["R2"])
            delta_rmse = float(row["RMSE"] - base["RMSE"])
            if feature_set == "All descriptors":
                interpretation = "Baseline; target labels were not imputed."
            elif model == "TOPSIS_ensemble" and abs(delta_r2) <= 0.05 and abs(delta_rmse) <= 0.10:
                interpretation = "Main performance is stable after removing high-missing descriptors."
            elif model == "TOPSIS_ensemble":
                interpretation = "Performance changes after removing high-missing descriptors; discuss imputation sensitivity."
            else:
                interpretation = "Sensitivity check for individual model."
            rows.append(
                {
                    "Feature_set": feature_set,
                    "Missingness_rule": "Target variables were never imputed; this check removes input descriptors with high WGS-domain missingness.",
                    "Removed_descriptor_sources": "" if feature_set == "All descriptors" else removed_source_text,
                    "Removed_engineered_feature_count": 0 if feature_set == "All descriptors" else len(removed_cols),
                    "Retained_engineered_feature_count": len(context["x_wgs"].columns) if feature_set == "All descriptors" else len(retained_cols),
                    "Model": model,
                    "R2": round(float(row["R2"]), 4),
                    "RMSE": round(float(row["RMSE"]), 4),
                    "MAE": round(float(row["MAE"]), 4),
                    "MAPE": round(float(row["MAPE"]), 4),
                    "Delta_R2_vs_all_descriptors": round(delta_r2, 4),
                    "Delta_RMSE_vs_all_descriptors": round(delta_rmse, 4),
                    "Interpretation": interpretation,
                }
            )
    return pd.DataFrame(rows)


def make_table4(context: dict, fig) -> pd.DataFrame:
    dop = context["dop"].copy()
    dop["Predicted E_ads (eV)"] = context["preds_dop"]["TOPSIS_ensemble"]
    dop["Residual (eV)"] = dop["Predicted E_ads (eV)"] - dop["E_ads_ev"]
    dop["Dopant"] = dop["Dopant"].replace(fig.ALIASES).fillna("None")
    dop["DFT-window score"] = np.exp(-((dop["E_ads_ev"] + 0.80) / 0.90) ** 2)
    dop["Predicted-window score"] = np.exp(-((dop["Predicted E_ads (eV)"] + 0.80) / 0.90) ** 2)
    rank = (
        dop.groupby("Dopant")
        .agg(
            Samples=("E_ads_ev", "size"),
            Mean_DFT_Eads_eV=("E_ads_ev", "mean"),
            Mean_pred_Eads_eV=("Predicted E_ads (eV)", "mean"),
            Transfer_MAE_eV=("Residual (eV)", lambda s: float(np.mean(np.abs(s)))),
            Mean_DFT_window_score=("DFT-window score", "mean"),
            Mean_pred_window_score=("Predicted-window score", "mean"),
        )
        .query("Dopant != 'None'")
        .sort_values("Mean_DFT_window_score", ascending=False)
        .reset_index()
    )
    for col in rank.columns:
        if col != "Dopant":
            rank[col] = rank[col].astype(float).round(4)
    return rank


def make_table4_metrics(context: dict, fig) -> pd.DataFrame:
    dop = context["dop"].copy()
    dop["Predicted E_ads (eV)"] = context["preds_dop"]["TOPSIS_ensemble"]
    m = fig.metrics(dop["E_ads_ev"], dop["Predicted E_ads (eV)"])
    rho, p = spearmanr(dop["E_ads_ev"], dop["Predicted E_ads (eV)"], nan_policy="omit")
    return pd.DataFrame(
        [
            {
                "Dataset": "dopants external domain",
                "Samples": int(len(dop)),
                "Transfer_R2": round(float(m["R2"]), 4),
                "Transfer_RMSE_eV": round(float(m["RMSE"]), 4),
                "Transfer_MAE_eV": round(float(m["MAE"]), 4),
                "Transfer_MAPE": round(float(m["MAPE"]), 4),
                "Spearman_rho": round(float(rho), 4),
                "Spearman_p": round(float(p), 6),
                "Interpretation": "Negative R2 indicates failed direct WGS-to-dopants transfer; use dopants labels as domain-shift diagnostic, not as evidence of reliable ML screening.",
            }
        ]
    )


def make_table3_dopant_summary(context: dict, fig) -> pd.DataFrame:
    metrics_df = make_table4_metrics(context, fig)
    rank = make_table4(context, fig)
    metric = metrics_df.iloc[0]
    rows = [
        {
            "Section": "Overall transfer diagnostic",
            "Item": "WGS model transferred to dopants",
            "Samples": int(metric["Samples"]),
            "Transfer_R2": metric["Transfer_R2"],
            "Transfer_RMSE_eV": metric["Transfer_RMSE_eV"],
            "Transfer_MAE_eV": metric["Transfer_MAE_eV"],
            "Spearman_rho": metric["Spearman_rho"],
            "Mean_DFT_Eads_eV": "",
            "Mean_pred_Eads_eV": "",
            "Dopant_MAE_eV": "",
            "Mean_DFT_window_score": "",
            "Mean_pred_window_score": "",
            "Interpretation": "Negative R2 indicates failed direct transfer; this is a domain-shift diagnostic, not evidence of reliable ML screening.",
        }
    ]
    concise_rank = rank[rank["Samples"] >= 3].head(8).copy()
    for _, row in concise_rank.iterrows():
        rows.append(
            {
                "Section": "DFT-labelled dopant trend",
                "Item": row["Dopant"],
                "Samples": int(row["Samples"]),
                "Transfer_R2": "",
                "Transfer_RMSE_eV": "",
                "Transfer_MAE_eV": "",
                "Spearman_rho": "",
                "Mean_DFT_Eads_eV": row["Mean_DFT_Eads_eV"],
                "Mean_pred_Eads_eV": row["Mean_pred_Eads_eV"],
                "Dopant_MAE_eV": row["Transfer_MAE_eV"],
                "Mean_DFT_window_score": row["Mean_DFT_window_score"],
                "Mean_pred_window_score": row["Mean_pred_window_score"],
                "Interpretation": "Ranked by DFT-window score using dopants-sheet DFT labels; n>=3 retained for main-text summary.",
            }
        )
    return pd.DataFrame(rows)


def make_table_s1(wgs_raw: pd.DataFrame, dop_raw: pd.DataFrame) -> pd.DataFrame:
    rows = []
    for sheet, df in [("WGS", wgs_raw), ("dopants", dop_raw)]:
        for col in df.columns:
            s = df[col].replace("-", np.nan)
            rows.append(
                {
                    "Dataset": sheet,
                    "Column": col,
                    "Valid count": int(s.notna().sum()),
                    "Missing count": int(s.isna().sum()),
                    "Missing rate (%)": round(s.isna().mean() * 100, 2),
                    "Unique values": int(s.nunique(dropna=True)),
                }
            )
    return pd.DataFrame(rows)


def make_table_s2(context: dict) -> tuple[pd.DataFrame, pd.DataFrame]:
    topsis = context["topsis"].reset_index(names="Model").round(5)
    models = [m for m in context["metrics"].index if m != "TOPSIS_ensemble"] + ["TOPSIS_ensemble"]
    y = context["y_test"].to_numpy()
    pmat = pd.DataFrame(np.ones((len(models), len(models))), index=models, columns=models)
    for a in models:
        for b in models:
            if a == b:
                pmat.loc[a, b] = np.nan
            else:
                try:
                    pmat.loc[a, b] = wilcoxon(np.abs(y - context["preds_test"][a]), np.abs(y - context["preds_test"][b])).pvalue
                except ValueError:
                    pmat.loc[a, b] = 1.0
    return topsis, pmat.round(6)


def make_table_s3(wgs: pd.DataFrame) -> pd.DataFrame:
    subset = wgs[wgs["E_activation_ev"].notna()].copy()
    out = (
        subset.groupby("Reaction")
        .agg(
            Samples=("E_activation_ev", "size"),
            Eads_mean_eV=("E_ads_ev", "mean"),
            Eads_std_eV=("E_ads_ev", "std"),
            Ea_mean_eV=("E_activation_ev", "mean"),
            Ea_std_eV=("E_activation_ev", "std"),
            Ereaction_mean_eV=("E_reaction_ev", "mean"),
            Ereaction_std_eV=("E_reaction_ev", "std"),
        )
        .reset_index()
    )
    for col in out.columns:
        if col != "Reaction":
            out[col] = out[col].astype(float).round(4)
    return out


def make_table_s3_combined(wgs: pd.DataFrame, context: dict, fig) -> pd.DataFrame:
    activation = make_table_s3(wgs)
    dopant_detail = make_table4(context, fig)
    cols = [
        "Section",
        "Group",
        "Samples",
        "Eads_mean_eV",
        "Eads_std_eV",
        "Ea_mean_eV",
        "Ea_std_eV",
        "Ereaction_mean_eV",
        "Ereaction_std_eV",
        "Mean_pred_Eads_eV",
        "Transfer_MAE_eV",
        "Mean_DFT_window_score",
        "Mean_pred_window_score",
        "Notes",
    ]
    rows = []
    for _, row in activation.iterrows():
        rows.append(
            {
                "Section": "Activation-barrier subset",
                "Group": row["Reaction"],
                "Samples": int(row["Samples"]),
                "Eads_mean_eV": row["Eads_mean_eV"],
                "Eads_std_eV": row["Eads_std_eV"],
                "Ea_mean_eV": row["Ea_mean_eV"],
                "Ea_std_eV": row["Ea_std_eV"],
                "Ereaction_mean_eV": row["Ereaction_mean_eV"],
                "Ereaction_std_eV": row["Ereaction_std_eV"],
                "Mean_pred_Eads_eV": "",
                "Transfer_MAE_eV": "",
                "Mean_DFT_window_score": "",
                "Mean_pred_window_score": "",
                "Notes": "Small-sample mechanistic subset; not used as the main prediction task.",
            }
        )
    for _, row in dopant_detail.iterrows():
        rows.append(
            {
                "Section": "Dopant-resolved transfer detail",
                "Group": row["Dopant"],
                "Samples": int(row["Samples"]),
                "Eads_mean_eV": row["Mean_DFT_Eads_eV"],
                "Eads_std_eV": "",
                "Ea_mean_eV": "",
                "Ea_std_eV": "",
                "Ereaction_mean_eV": "",
                "Ereaction_std_eV": "",
                "Mean_pred_Eads_eV": row["Mean_pred_Eads_eV"],
                "Transfer_MAE_eV": row["Transfer_MAE_eV"],
                "Mean_DFT_window_score": row["Mean_DFT_window_score"],
                "Mean_pred_window_score": row["Mean_pred_window_score"],
                "Notes": "Detailed dopant statistics; main text reports only a concise n>=3 trend summary.",
            }
        )
    return pd.DataFrame(rows, columns=cols)


def style_workbook(path: Path) -> None:
    wb = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="24435C")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2EC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.sheet_view.showGridLines = False
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = border
        for col_idx, col in enumerate(ws.columns, start=1):
            max_len = 10
            for cell in col:
                value = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, min(len(value), 55))
            ws.column_dimensions[get_column_letter(col_idx)].width = max(10, min(max_len + 2, 42))
    wb.save(path)


def write_table_notes() -> None:
    text = """# 表格说明

本文件夹包含按 CEJ 风格压缩后的正文表和补充表。

- `Paper_Tables_MXene_Eads.xlsx`：所有表格的总工作簿。
- `Table1_*.csv` 至 `Table3_*.csv`：正文表。
- `TableS1_*.csv` 至 `TableS4_*.csv`：补充表。
- `generate_paper_tables.py`：表格生成代码。
- `generate_paper_tables.ipynb`：Jupyter 可运行版本。

正文建议只放 3 张表：

- `Table1_Label_Availability.csv`：数据标签可用性与任务定义。
- `Table2_WGS_Model_Performance.csv`：WGS 域 hold-out 测试集模型性能。
- `Table3_Dopant_Domain_Transfer_and_DFT_Trend.csv`：dopants 外部域迁移诊断与 DFT 标注掺杂趋势摘要。

补充材料建议放 4 组表：

- `TableS1_Descriptor_Definitions.csv`：descriptor definitions、单位、缺失率和预处理策略。
- `TableS2_TOPSIS_Weights.csv` 与 `TableS2_Wilcoxon_P.csv`：TOPSIS 权重、closeness scores 和 Wilcoxon p values。
- `TableS3_Activation_and_Dopant_Details.csv`：62 条活化能子集统计与完整 dopant-resolved 明细。
- `TableS4_Missingness_Sensitivity.csv`：保留全部描述符 vs 移除 WGS 域缺失率 >=50% 描述符后的模型性能对照，用于回应缺失值填充风险。

注意：目标变量不做填充；缺失值填充只用于输入描述符，并在 `TableS4_Missingness_Sensitivity.csv` 中报告高缺失描述符剔除后的性能变化。当前 WGS 模型直接迁移到 dopants 的 R2 为负，因此正文 Table 3 不能用于支持“高精度外部验证”或“模型可靠筛选掺杂元素”；它用于报告迁移失败，并基于 dopants 表已有 DFT Eads 标签给出吸附窗口趋势。
"""
    (OUT / "表格说明.md").write_text(text, encoding="utf-8")


def write_notebook() -> None:
    nb = {
        "cells": [
            {
                "cell_type": "markdown",
                "metadata": {},
                "source": [
                    "# MXene adsorption-energy paper tables\n",
                    "\n",
                    "Run the next cell to regenerate all paper tables from the Excel dataset and the same modeling workflow used for the figures.\n",
                ],
            },
            {
                "cell_type": "code",
                "execution_count": None,
                "metadata": {},
                "outputs": [],
                "source": [
                    "from pathlib import Path\n",
                    "import runpy\n",
                    "\n",
                    "script = Path('generate_paper_tables.py')\n",
                    "runpy.run_path(str(script), run_name='__main__')\n",
                ],
            },
        ],
        "metadata": {
            "kernelspec": {"display_name": "Python 3", "language": "python", "name": "python3"},
            "language_info": {"name": "python", "version": "3.x"},
        },
        "nbformat": 4,
        "nbformat_minor": 5,
    }
    (OUT / "generate_paper_tables.ipynb").write_text(json.dumps(nb, ensure_ascii=False, indent=2), encoding="utf-8")


def main() -> None:
    fig = load_figure_module()
    fig.configure_style()
    wgs, dop = fig.load_data()
    context = fig.prepare_training(wgs, dop)
    wgs_raw = pd.read_excel(fig.DATA_PATH, sheet_name="WGS")
    dop_raw = pd.read_excel(fig.DATA_PATH, sheet_name="dopants")

    topsis_weights, wilcoxon_p = make_table_s2(context)
    tables = {
        "Table1_Label_Availability": make_table1(wgs_raw, dop_raw),
        "Table2_WGS_Model_Performance": make_table3(context),
        "Table3_Dopant_Domain_Transfer_and_DFT_Trend": make_table3_dopant_summary(context, fig),
        "TableS1_Descriptor_Definitions": make_table2(wgs_raw.replace("-", np.nan), dop_raw.replace("-", np.nan)),
        "TableS2_TOPSIS_Weights": topsis_weights,
        "TableS2_Wilcoxon_P": wilcoxon_p.reset_index(names="Model"),
        "TableS3_Activation_and_Dopant_Details": make_table_s3_combined(wgs, context, fig),
        "TableS4_Missingness_Sensitivity": make_table_s1b_missingness_sensitivity(context, fig, wgs_raw.replace("-", np.nan)),
    }

    for old in OUT.glob("Table*.csv"):
        old.unlink()
    for name, df in tables.items():
        df.to_csv(OUT / f"{name}.csv", index=False, encoding="utf-8-sig")

    workbook = OUT / "Paper_Tables_MXene_Eads.xlsx"
    with pd.ExcelWriter(workbook, engine="openpyxl") as writer:
        for name, df in tables.items():
            sheet = name[:31]
            df.to_excel(writer, sheet_name=sheet, index=False)
    style_workbook(workbook)
    write_table_notes()
    write_notebook()
    print(f"Saved tables to: {OUT}")


if __name__ == "__main__":
    main()
